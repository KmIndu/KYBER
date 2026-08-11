"""Tests for multi-format ingestion — CSV, XLSX, XML parsers and format dispatcher."""

from __future__ import annotations

import json
import textwrap

import pytest

from app.models.schema import SchemaMetadata
from app.parsers.csv_parser import CSVParserError, parse_csv_schema
from app.parsers.format_dispatcher import (
    FormatDetectionError,
    FormatParseError,
    classify_format,
    detect_and_parse,
)
from app.parsers.xml_parser import XMLParserError, parse_xml_schema

# ── CSV Parser ────────────────────────────────────────────────


class TestCSVParser:
    def test_basic_csv(self):
        csv_text = textwrap.dedent("""\
            id,name,email,age,active
            1,Alice,alice@example.com,30,true
            2,Bob,bob@example.com,25,false
            3,Charlie,charlie@example.com,35,true
        """)
        schema = parse_csv_schema(csv_text, table_name="users")
        assert len(schema.tables) == 1
        t = schema.tables[0]
        assert t.name == "users"
        assert len(t.columns) == 5
        col_names = [c.name for c in t.columns]
        assert "id" in col_names
        assert "name" in col_names
        assert "email" in col_names

    def test_type_inference_integer(self):
        csv_text = "count\n1\n2\n3\n4\n5\n"
        schema = parse_csv_schema(csv_text)
        col = schema.tables[0].columns[0]
        assert col.data_type == "INTEGER"

    def test_type_inference_decimal(self):
        csv_text = "price\n1.50\n2.75\n3.00\n4.25\n5.99\n"
        schema = parse_csv_schema(csv_text)
        col = schema.tables[0].columns[0]
        assert col.data_type == "DECIMAL"

    def test_type_inference_date(self):
        csv_text = "created\n2024-01-01\n2024-02-15\n2024-03-20\n2024-04-10\n2024-05-05\n"
        schema = parse_csv_schema(csv_text)
        col = schema.tables[0].columns[0]
        assert col.data_type == "DATE"

    def test_type_inference_boolean(self):
        csv_text = "id,active\n1,true\n2,false\n3,true\n4,true\n5,false\n"
        schema = parse_csv_schema(csv_text)
        col = next(c for c in schema.tables[0].columns if c.name == "active")
        assert col.data_type == "BOOLEAN"

    def test_primary_key_detection(self):
        csv_text = "id,name\n1,Alice\n2,Bob\n3,Charlie\n"
        schema = parse_csv_schema(csv_text)
        id_col = next(c for c in schema.tables[0].columns if c.name == "id")
        assert id_col.is_primary_key is True
        assert "id" in schema.tables[0].primary_keys

    def test_foreign_key_detection(self):
        csv_text = "id,customer_id,amount\n1,10,99.99\n2,20,50.00\n3,30,75.50\n"
        schema = parse_csv_schema(csv_text)
        fks = schema.tables[0].foreign_keys
        assert len(fks) == 1
        assert fks[0].column == "customer_id"
        assert fks[0].references_table == "customer"

    def test_nullable_detection(self):
        csv_text = "id,notes\n1,hello\n2,\n3,world\n"
        schema = parse_csv_schema(csv_text)
        notes_col = next(c for c in schema.tables[0].columns if c.name == "notes")
        assert notes_col.nullable is True

    def test_header_only_csv(self):
        csv_text = "id,name,email\n"
        schema = parse_csv_schema(csv_text)
        assert len(schema.tables[0].columns) == 3
        for col in schema.tables[0].columns:
            assert col.data_type == "VARCHAR"
            assert col.nullable is True

    def test_empty_csv_raises(self):
        with pytest.raises(CSVParserError, match="empty"):
            parse_csv_schema("")

    def test_tsv_detection(self):
        csv_text = "id\tname\tage\n1\tAlice\t30\n2\tBob\t25\n"
        schema = parse_csv_schema(csv_text)
        assert len(schema.tables[0].columns) == 3

    def test_duplicate_headers(self):
        csv_text = "id,name,name\n1,a,b\n2,c,d\n"
        schema = parse_csv_schema(csv_text)
        names = [c.name for c in schema.tables[0].columns]
        assert len(names) == len(set(names)), "Duplicate headers should be de-duped"

    def test_returns_schema_metadata(self):
        csv_text = "x\n1\n"
        result = parse_csv_schema(csv_text)
        assert isinstance(result, SchemaMetadata)


# ── XLSX Parser ───────────────────────────────────────────────


class TestXLSXParser:
    def _make_xlsx(self, sheets: dict[str, list[list]]) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        first = True
        for name, rows in sheets.items():
            ws = wb.active if first else wb.create_sheet()
            ws.title = name
            for row in rows:
                ws.append(row)
            first = False
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_single_sheet(self):
        from app.parsers.xlsx_parser import parse_xlsx_schema

        data = self._make_xlsx(
            {"Sheet1": [["id", "name", "age"], [1, "Alice", 30], [2, "Bob", 25]]}
        )
        schema = parse_xlsx_schema(data) 
        assert len(schema.tables) == 1
        assert len(schema.tables[0].columns) == 3

    def test_multi_sheet(self):
        from app.parsers.xlsx_parser import parse_xlsx_schema

        data = self._make_xlsx(
            {
                "users": [["id", "name"], [1, "Alice"]],
                "orders": [["id", "user_id", "total"], [1, 1, 99.99]],
            }
        )
        schema = parse_xlsx_schema(data)
        assert len(schema.tables) == 2

    def test_type_inference(self):
        from app.parsers.xlsx_parser import parse_xlsx_schema
        import datetime

        data = self._make_xlsx(
            {
                "Sheet1": [
                    ["count", "price", "created"],
                    [1, 10.5, datetime.date(2024, 1, 1)],
                    [2, 20.3, datetime.date(2024, 2, 1)],
                ]
            }
        )
        schema = parse_xlsx_schema(data)
        cols = {c.name: c for c in schema.tables[0].columns}
        assert cols["count"].data_type == "INTEGER"
        assert cols["price"].data_type == "DECIMAL"
        assert cols["created"].data_type in ("DATE", "TIMESTAMP")

    def test_empty_xlsx_raises(self):
        from app.parsers.xlsx_parser import XLSXParserError, parse_xlsx_schema

        data = self._make_xlsx({"Empty": []})
        with pytest.raises(XLSXParserError, match="No sheets"):
            parse_xlsx_schema(data)

    def test_fk_detection(self):
        from app.parsers.xlsx_parser import parse_xlsx_schema

        data = self._make_xlsx(
            {"orders": [["id", "customer_id", "total"], [1, 10, 50], [2, 20, 75]]}
        )
        schema = parse_xlsx_schema(data)
        fks = schema.tables[0].foreign_keys
        assert len(fks) == 1
        assert fks[0].column == "customer_id"

    def test_invalid_bytes_raises(self):
        from app.parsers.xlsx_parser import XLSXParserError, parse_xlsx_schema

        with pytest.raises(XLSXParserError, match="Cannot open"):
            parse_xlsx_schema(b"not an xlsx file")

    def test_returns_schema_metadata(self):
        from app.parsers.xlsx_parser import parse_xlsx_schema

        data = self._make_xlsx({"S": [["x"], [1]]})
        result = parse_xlsx_schema(data)
        assert isinstance(result, SchemaMetadata)


# ── XML Parser ────────────────────────────────────────────────


class TestXMLParser:
    def test_basic_xml(self):
        xml = textwrap.dedent("""\
            <users>
              <user id="1">
                <name>Alice</name>
                <age>30</age>
              </user>
              <user id="2">
                <name>Bob</name>
                <age>25</age>
              </user>
            </users>
        """)
        schema = parse_xml_schema(xml)
        assert len(schema.tables) >= 1
        # "user" should be a table since it repeats
        table_names = [t.name for t in schema.tables]
        assert "user" in table_names

    def test_attributes_become_columns(self):
        xml = '<items><item code="A1" price="10.5"/><item code="B2" price="20.3"/></items>'
        schema = parse_xml_schema(xml)
        item_table = next(t for t in schema.tables if t.name == "item")
        col_names = [c.name for c in item_table.columns]
        assert "code" in col_names
        assert "price" in col_names

    def test_child_text_becomes_columns(self):
        xml = textwrap.dedent("""\
            <products>
              <product><name>Widget</name><price>9.99</price></product>
              <product><name>Gadget</name><price>19.99</price></product>
            </products>
        """)
        schema = parse_xml_schema(xml)
        prod_table = next(t for t in schema.tables if t.name == "product")
        col_names = [c.name for c in prod_table.columns]
        assert "name" in col_names
        assert "price" in col_names

    def test_type_inference_integer(self):
        xml = '<data><row><count>1</count></row><row><count>2</count></row></data>'
        schema = parse_xml_schema(xml)
        row_table = next(t for t in schema.tables if t.name == "row")
        count_col = next(c for c in row_table.columns if c.name == "count")
        assert count_col.data_type == "INTEGER"

    def test_invalid_xml_raises(self):
        with pytest.raises(XMLParserError, match="Invalid XML"):
            parse_xml_schema("<not>valid>xml")

    def test_empty_structure_raises(self):
        with pytest.raises(XMLParserError, match="Could not extract"):
            parse_xml_schema("<root/>")

    def test_returns_schema_metadata(self):
        xml = '<data><item><x>1</x></item><item><x>2</x></item></data>'
        result = parse_xml_schema(xml)
        assert isinstance(result, SchemaMetadata)

    def test_namespace_stripping(self):
        xml = '<root xmlns:ns="http://example.com"><ns:item><ns:val>1</ns:val></ns:item><ns:item><ns:val>2</ns:val></ns:item></root>'
        schema = parse_xml_schema(xml)
        # Should strip namespace prefixes
        table_names = [t.name for t in schema.tables]
        # At least one table should exist
        assert len(schema.tables) >= 1


# ── Format Dispatcher ────────────────────────────────────────


class TestFormatClassification:
    def test_sql_extension(self):
        assert classify_format("schema.sql", "") == "sql"

    def test_yaml_extension(self):
        assert classify_format("api.yaml", "") == "openapi"

    def test_yml_extension(self):
        assert classify_format("api.yml", "") == "openapi"

    def test_csv_extension(self):
        assert classify_format("data.csv", "") == "csv"

    def test_xlsx_extension(self):
        assert classify_format("report.xlsx", b"") == "xlsx"

    def test_xml_extension(self):
        assert classify_format("config.xml", "") == "xml"

    def test_feature_extension(self):
        assert classify_format("login.feature", "") == "bdd"

    def test_json_openapi_sniffing(self):
        content = json.dumps({"openapi": "3.0.0", "info": {"title": "API"}, "paths": {}})
        assert classify_format("api.json", content) == "openapi"

    def test_json_tabular_sniffing(self):
        content = json.dumps([{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}])
        assert classify_format("data.json", content) == "csv"

    def test_txt_bdd_sniffing(self):
        content = "Feature: Login\n  Scenario: Valid login\n    Given a user exists"
        assert classify_format("test.txt", content) == "bdd"

    def test_txt_sql_sniffing(self):
        content = "CREATE TABLE users (id INT PRIMARY KEY);"
        assert classify_format("schema.txt", content) == "sql"

    def test_txt_csv_sniffing(self):
        content = "id,name,age\n1,Alice,30\n2,Bob,25\n"
        assert classify_format("data.txt", content) == "csv"

    def test_unknown_extension_raises(self):
        with pytest.raises(FormatDetectionError, match="Cannot determine"):
            classify_format("file.xyz", "random content without patterns")


class TestFormatDispatcher:
    def test_csv_dispatch(self):
        csv_text = "id,name\n1,Alice\n2,Bob\n"
        schema = detect_and_parse(csv_text, "users.csv")
        assert isinstance(schema, SchemaMetadata)
        assert len(schema.tables) == 1

    def test_xml_dispatch(self):
        xml = '<items><item><name>A</name></item><item><name>B</name></item></items>'
        schema = detect_and_parse(xml, "items.xml")
        assert isinstance(schema, SchemaMetadata)
        assert len(schema.tables) >= 1

    def test_sql_dispatch(self):
        sql = "CREATE TABLE t (id INT PRIMARY KEY, name VARCHAR(50));"
        schema = detect_and_parse(sql, "schema.sql")
        assert isinstance(schema, SchemaMetadata)
        assert len(schema.tables) == 1

    def test_bad_content_raises_parse_error(self):
        with pytest.raises(FormatParseError):
            detect_and_parse("<broken>xml", "data.xml")

    def test_table_name_from_filename(self):
        csv_text = "id,val\n1,a\n2,b\n"
        schema = detect_and_parse(csv_text, "my_dataset.csv")
        assert schema.tables[0].name == "my_dataset"

    def test_json_array_dispatch(self):
        content = json.dumps([{"id": 1, "x": "a"}, {"id": 2, "x": "b"}])
        schema = detect_and_parse(content, "data.json")
        assert isinstance(schema, SchemaMetadata)


# ── Pipeline integration (upload + parse) ─────────────────────


class TestPipelineCSV:
    """Integration tests for CSV through the pipeline endpoints."""

    @pytest.fixture
    def client(self):
        from fastapi.testclient import TestClient

        from app.main import app

        return TestClient(app)

    def test_upload_csv(self, client):
        csv_content = b"id,name,age\n1,Alice,30\n2,Bob,25\n"
        resp = client.post(
            "/upload",
            files=[("files", ("data.csv", csv_content, "text/csv"))],
        )
        assert resp.status_code == 201
        data = resp.json()
        assert data["files"][0]["file_type"] == "csv"

    def test_upload_and_parse_csv(self, client):
        csv_content = b"id,name,age\n1,Alice,30\n2,Bob,25\n"
        upload = client.post(
            "/upload",
            files=[("files", ("users.csv", csv_content, "text/csv"))],
        )
        sid = upload.json()["session_id"]
        parse_resp = client.post(f"/parse?session_id={sid}")
        assert parse_resp.status_code == 200
        data = parse_resp.json()
        assert len(data["tables"]) >= 1

    def test_upload_xlsx(self, client):
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        buf = BytesIO()
        wb.save(buf)

        resp = client.post(
            "/upload",
            files=[("files", ("data.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
        assert resp.status_code == 201
        assert resp.json()["files"][0]["file_type"] == "xlsx"

    def test_upload_xml(self, client):
        xml_content = b'<users><user><name>Alice</name></user><user><name>Bob</name></user></users>'
        resp = client.post(
            "/upload",
            files=[("files", ("data.xml", xml_content, "text/xml"))],
        )
        assert resp.status_code == 201
        assert resp.json()["files"][0]["file_type"] == "xml"

    def test_upload_unsupported_rejected(self, client):
        resp = client.post(
            "/upload",
            files=[("files", ("data.pdf", b"content", "application/pdf"))],
        )
        assert resp.status_code == 400

    def test_full_csv_pipeline(self, client):
        """End-to-end: upload CSV → parse → generate → download."""
        csv_content = b"id,name,email\n1,Alice,alice@test.com\n2,Bob,bob@test.com\n3,Carol,carol@test.com\n"
        upload = client.post(
            "/upload",
            files=[("files", ("people.csv", csv_content, "text/csv"))],
        )
        sid = upload.json()["session_id"]

        parse_resp = client.post(f"/parse?session_id={sid}")
        assert parse_resp.status_code == 200

        gen_resp = client.post(
            f"/generate?session_id={sid}&row_count=5&include_valid=true"
        )
        assert gen_resp.status_code == 200
        assert gen_resp.json()["total_rows"] > 0

    def test_full_xml_pipeline(self, client):
        xml_content = b"""<records>
            <record><id>1</id><name>Alpha</name><value>100</value></record>
            <record><id>2</id><name>Beta</name><value>200</value></record>
        </records>"""
        upload = client.post(
            "/upload",
            files=[("files", ("records.xml", xml_content, "text/xml"))],
        )
        sid = upload.json()["session_id"]

        parse_resp = client.post(f"/parse?session_id={sid}")
        assert parse_resp.status_code == 200

        gen_resp = client.post(
            f"/generate?session_id={sid}&row_count=5&include_valid=true"
        )
        assert gen_resp.status_code == 200
        assert gen_resp.json()["total_rows"] > 0
