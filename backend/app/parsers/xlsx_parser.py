"""XLSX parser — infers schema metadata from Excel workbooks.

Each worksheet becomes a separate TableMetadata entry.
Headers are read from the first row; data types are inferred
by sampling cell values and types.
"""

from __future__ import annotations

import logging
import re
from collections import Counter
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.models.schema import (
    ColumnMetadata,
    ForeignKeyMetadata,
    SchemaMetadata,
    TableMetadata,
)

logger = logging.getLogger(__name__)

_MAX_SAMPLE_ROWS = 500


class XLSXParserError(Exception):
    """Raised when XLSX parsing fails."""


def parse_xlsx_schema(
    raw_bytes: bytes,
    default_table_name: str = "imported_table",
) -> SchemaMetadata:
    """Infer SQL schema from an Excel workbook.

    Each sheet with a header row becomes a separate table.
    """
    try:
        wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    except (InvalidFileException, Exception) as exc:
        raise XLSXParserError(f"Cannot open XLSX file: {exc}") from exc

    tables: list[TableMetadata] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows()

        # Read header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        headers = [_clean_header(cell.value) for cell in header_row]
        if not headers or all(h == "" for h in headers):
            continue

        # De-duplicate headers
        seen: dict[str, int] = {}
        clean_headers: list[str] = []
        for h in headers:
            if not h:
                h = "column"
            if h in seen:
                seen[h] += 1
                h = f"{h}_{seen[h]}"
            else:
                seen[h] = 0
            clean_headers.append(h)
        headers = clean_headers

        # Sample data rows
        col_values: list[list] = [[] for _ in headers]
        n_rows = 0
        for row in rows_iter:
            if n_rows >= _MAX_SAMPLE_ROWS:
                break
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    col_values[idx].append(cell.value)
            n_rows += 1

        # Build columns
        columns: list[ColumnMetadata] = []
        primary_keys: list[str] = []

        for col_idx, col_name in enumerate(headers):
            raw_vals = col_values[col_idx] if col_idx < len(col_values) else []
            non_null = [v for v in raw_vals if v is not None and str(v).strip() != ""]
            has_nulls = len(non_null) < n_rows if n_rows > 0 else True
            data_type = _infer_type_from_cells(non_null)

            unique_strs = {str(v) for v in non_null}
            is_unique = len(unique_strs) == len(non_null) and len(non_null) == n_rows and n_rows > 0

            is_pk = False
            if is_unique and not has_nulls and _looks_like_id(col_name):
                is_pk = True
                primary_keys.append(col_name)

            columns.append(
                ColumnMetadata(
                    name=col_name,
                    data_type=data_type,
                    nullable=has_nulls,
                    is_primary_key=is_pk,
                    is_unique=is_unique and not is_pk,
                )
            )

        # FK heuristics
        fks: list[ForeignKeyMetadata] = []
        for col in columns:
            if col.name.endswith("_id") and not col.is_primary_key:
                fks.append(
                    ForeignKeyMetadata(
                        column=col.name,
                        references_table=col.name[:-3],
                        references_column="id",
                    )
                )

        table_name = _clean_header(sheet_name) or default_table_name
        if len(wb.sheetnames) == 1:
            table_name = default_table_name

        tables.append(
            TableMetadata(
                name=table_name,
                columns=columns,
                primary_keys=primary_keys,
                foreign_keys=fks,
            )
        )

    wb.close()

    if not tables:
        raise XLSXParserError("No sheets with header rows found in XLSX file")

    logger.info(
        "XLSX parsed — %d table(s) extracted",
        len(tables),
        extra={"stage": "parsing", "event": "xlsx_parsed"},
    )

    return SchemaMetadata(tables=tables)


# ── Helpers ───────────────────────────────────────────────────

_HEADER_CLEAN_RE = re.compile(r"[^\w]+")


def _clean_header(raw) -> str:
    if raw is None:
        return ""
    cleaned = str(raw).strip()
    cleaned = _HEADER_CLEAN_RE.sub("_", cleaned).strip("_").lower()
    return cleaned


_ID_PATTERNS = re.compile(r"(?:^id$|^pk$|^key$|^uuid$)", re.IGNORECASE)


def _looks_like_id(col_name: str) -> bool:
    return bool(_ID_PATTERNS.search(col_name))


import datetime


def _infer_type_from_cells(values: list) -> str:
    """Infer a SQL data type from openpyxl cell values (already typed)."""
    if not values:
        return "VARCHAR"

    type_counts: Counter[str] = Counter()

    for v in values:
        if isinstance(v, bool):
            type_counts["BOOLEAN"] += 1
        elif isinstance(v, int):
            if -(2**31) <= v <= 2**31 - 1:
                type_counts["INTEGER"] += 1
            else:
                type_counts["BIGINT"] += 1
        elif isinstance(v, float):
            type_counts["DECIMAL"] += 1
        elif isinstance(v, datetime.datetime):
            type_counts["TIMESTAMP"] += 1
        elif isinstance(v, datetime.date):
            type_counts["DATE"] += 1
        elif isinstance(v, datetime.time):
            type_counts["TIME"] += 1
        else:
            type_counts["VARCHAR"] += 1

    if not type_counts:
        return "VARCHAR"

    dominant, dominant_count = type_counts.most_common(1)[0]
    if dominant_count / len(values) >= 0.8:
        return dominant

    return "VARCHAR"
